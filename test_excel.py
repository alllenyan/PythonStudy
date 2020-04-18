import openpyxl
import datetime

# wb = openpyxl.Workbook()
# ws = wb.active()
# wa = ws.title
# ws['A1'] = 520
# ws.append([1,2,3]) #A2,B2,C3赋值
# ws['A3'] = datetime.datetime.now()
# wb.save('demo.xlxs')

a = 2 ** 0.5
print(a)
print(int(a))

print('''这是你%s月的账单，一共消费%f元，还剩%f元''' % ('二', 20.0, 15.0))

# 列表list学习
classmate = ['james', 'michle', 'sink']
print(classmate[0])
print(classmate[1])
print(classmate[2])

# len()方法获取列表的长度
print(len(classmate))

# appen()方法和insert()方法，可以往列表中加入数据，append()方法会将数据加入到列表的最后一行，insert()
# 方法会将数据加入到指定位置
print(classmate.append('amada'))
print(classmate.insert(1,'who'))


#pop()方法可以将列表中的数据删除，删除的数据为列表末尾的数据
# 如果要删除指定位置的数据，在pop()函数中指定索引即可
print(classmate.pop())

#元组,元组定义好之后不可更改
classmate2 = ('james', 'michle', 'sink')
print(classmate2)

d = {'Michael': 95, 'Bob': 75, 'Tracy': 85}

s = set([1, 2, 3])